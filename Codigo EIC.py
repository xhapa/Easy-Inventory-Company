from tkinter import *
import tkinter.font as tkFont  # importar libreria para manejo fuentes de texto
from time import strftime
from datetime import date,datetime
from openpyxl import Workbook,cell,load_workbook
import os
import random
import xlrd
import turtle
import time
#-----------------------------------------------------------------------------------------------------
"""
Funcion que importa el inventario de la
primera ventana del excel prueba.xlsx
lo transforma en una lista para poder
adaptarmo mas facilmente a un diccionario
que se retorna
"""
def importar():
    #intenta abrir el archivo
    try:
        wb2 = xlrd.open_workbook('prueba.xlsx')
        #Asigna la hoja en la posicion 0 a sheet
        sheet = wb2.sheet_by_index(0)
        #intenta hallar el numero de columnas en uso
        #midiendo la longitud de datos de la fila 0
        try:
            dir_cantidad = len(sheet.row_values(0))
        #si no puede la asigna a 1
        except:
            dir_cantidad = 1
        # intenta hallar el numero de filas en uso
        # midiendo la longitud de datos de la columna 0
        try:
            dir_cantidad2 = len(sheet.col_values(0))
        # si no puede la asigna a 1
        except:
            dir_cantidad2 = 1

        lista = []
        #Variable de control
        k = 1
        #un ciclo que recorre la longitud de las filas
        for i in range(0, dir_cantidad2 - 1):
            #añade una lista a la lista
            lista.append([])
            #un cilco que recorre las columnas
            for y in range(0, dir_cantidad):
                #asigna el valor de la celda en (k,y) a connection
                connection = sheet.cell(k, y).value
                #le agrega el valor de connection a la lista
                lista[i].append(connection)
            #suma al contador
            k += 1
        #diccionario que contendra el inventario
        dic = {}
        #ciclo que recorre la lista con el inventario
        for j in range(len(lista)):
            #si la seccion no es una llave de el diccionario
            if lista[j][0] not in dic.keys():
                #asigna la seccion como llave del diccionario que a su vez tambien es un diccionario
                dic[lista[j][0]] = {}
                #asigna al diccionario en la llave de la seccion y llave del id los valores del inventario
                dic[lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3], lista[j][4], lista[j][5], lista[j][6],lista[j][7]]
            #si la seccion ya es una llave del diccionario
            else:
                #asigna al diccionario en la llave de la seccion y la llave del id una lista vacia
                dic[lista[j][0]][lista[j][1]] = []
                ##asigna al diccionario en la llave de la seccion y la llave del id los valores del inventario
                dic[lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3], lista[j][4], lista[j][5], lista[j][6],lista[j][7]]
        #retorna el diccionario
        return dic
    #si el archivo no existe
    except:
        print("el archivo no existe")

"""
Funcion que importa el inventario de gastos
de la segunda ventana del excel prueba.xlsx
lo transforma en una lista para poder
adaptarmo mas facilmente a un diccionario
que se retorna
"""
def importar2():
    try:
        wb2 = xlrd.open_workbook('prueba.xlsx')
        sheet = wb2.sheet_by_index(1)
        try:
            dir_cantidad = len(sheet.row_values(0))
        except:
            dir_cantidad = 1
        try:
            dir_cantidad2 = len(sheet.col_values(0))
        except:
            dir_cantidad2 = 1

        lista = []
        k = 1
        for i in range(0, dir_cantidad2 - 1):
            lista.append([])
            for y in range(0, dir_cantidad):
                connection = sheet.cell(k, y).value
                lista[i].append(str(connection))
            k += 1
        dic = {}
        for j in range(len(lista)):
            if lista[j][7] not in dic.keys():
                dic[lista[j][7]] = {}
                dic[lista[j][7]][lista[j][0]]={}
                dic[lista[j][7]][lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3], lista[j][4], lista[j][5],lista[j][6]]
            else:
                if(lista[j][0] not in dic[lista[j][7]].keys()):
                    dic[lista[j][7]][lista[j][0]]={}
                    dic[lista[j][7]][lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3], lista[j][4], lista[j][5],lista[j][6]]
                else:
                    dic[lista[j][7]][lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3], lista[j][4], lista[j][5],lista[j][6]]
        return dic
    except:
        print("el archivo no existe")

"""
Funcion que importa el inventario de ganancias
de la tercera ventana del excel prueba.xlsx
lo transforma en una lista para poder
adaptarmo mas facilmente a un diccionario
que se retorna
"""
def importar3():
    try:
        wb2 = xlrd.open_workbook('prueba.xlsx')
        sheet = wb2.sheet_by_index(2)
        try:
            dir_cantidad = len(sheet.row_values(0))
        except:
            dir_cantidad = 1
        try:
            dir_cantidad2 = len(sheet.col_values(0))
        except:
            dir_cantidad2 = 1

        lista = []
        k = 1
        for i in range(0, dir_cantidad2 - 1):
            lista.append([])
            for y in range(0, dir_cantidad):
                connection = sheet.cell(k, y).value
                lista[i].append(str(connection))
            k += 1
        dic = {}
        for j in range(len(lista)):
            if lista[j][9] not in dic.keys():
                dic[lista[j][9]] = {}
                dic[lista[j][9]][lista[j][8]]={}
                dic[lista[j][9]][lista[j][8]][lista[j][0]] = {}
                dic[lista[j][9]][lista[j][8]][lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3], lista[j][4], lista[j][5],lista[j][6],lista[j][7]]
            else:
                if(lista[j][8] not in dic[lista[j][9]].keys()):
                    dic[lista[j][9]][lista[j][8]] = {}
                    dic[lista[j][9]][lista[j][8]][lista[j][0]] = {}
                    dic[lista[j][9]][lista[j][8]][lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3], lista[j][4],lista[j][5], lista[j][6], lista[j][7]]
                else:
                    if(lista[j][0] not in dic[lista[j][9]][lista[j][8]].keys()):
                        dic[lista[j][9]][lista[j][8]][lista[j][0]] = {}
                        dic[lista[j][9]][lista[j][8]][lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3],lista[j][4], lista[j][5],lista[j][6], lista[j][7]]
                    else:
                        dic[lista[j][9]][lista[j][8]][lista[j][0]][lista[j][1]] = [lista[j][2], lista[j][3],lista[j][4], lista[j][5],lista[j][6], lista[j][7]]
        return dic
    except:
        print("el archivo no existe")
#-----------------------------------------------------------------------------------------------------
"""
Funcion que mediante la libreria ctypes
obtiene el tamaño de la pantalla que
luego se usara para acomodar widgets
y asignarles un tamaño
retorna el ancho y el alto de la pantalla
"""
def tam_pantalla():
    import ctypes
    user32 = ctypes.windll.user32
    user32.SetProcessDPIAware()
    ancho, alto = user32.GetSystemMetrics(0), user32.GetSystemMetrics(1)
    return ancho,alto
"""
Listas que se usaran para
ir almacenando las direcciones
de las imagenes para que 
python no pierda la referencia
"""
imgs=[]
img=[]

"""
Funcion que asinga un tipo de fuente para
ser usado en algunas ventanas para algunos
widgets
retorna la fuente
"""
def fuente_salida():
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    fSalida = tkFont.Font(family="ALGERIAN", size=int(alt / 25))  # Cambiar tipo de fuente y tamaño
    return fSalida
"asigna una variable para ser usada en el reloj"
time1 = ''
def menu():
    """
    Funcion que configura el comportamiento
    y el tiempo de actualizacion del reloj
    """
    def reloj():
        global time1
        #formato del relok
        time2 = time.strftime('%H:%M:%S')
        #compara los estados de variables de tiempo
        if time2 != time1:
            time1 = time2
            #configura el reloj a una nueva hora
            clock.configure(text=time2)
        #tiempo actualizacion
        clock.after(500, reloj)
    anc,alt=tam_pantalla() #variables tamaño de pantalla
    window = Tk()  #Crear Ventana
    q = 0 #variable control para usuario
    """
    Si se inicia el programa
    crear la variable de control
    userAct y setearna a __
    """
    if (q == 0):
        userAct = StringVar(value="__")
        q = 1
    #nombres de las imagenes
    nomImgs = [["ingresar.png", "inventario.png", "gastos.png"],["venta.png", "ganancias.png", "salir.png"]]
    #configurar la ventana a pantalla completa
    window.attributes('-fullscreen',True)
    #asignar a atras el tipo de fuente que retorna fuente_salida
    atras = fuente_salida()
    #asignar un tipo de fuente diferente para los titulos
    fontStyle = tkFont.Font(family="Cooper Black", size=int(alt/20))  #Cambiar tipo de fuente y tamaño
    #lista con los textos de las opciones del menu
    opc=[["Ingreso Prod","Inventario","Gastos"],["Venta Prod","Rentabilidad","Salir"]]  #lista de opciones
    #lista con las funciones de cada boton
    funciones = [[subM_ingres,inventario,gastos],[salida_producto,ganancias,window.destroy]]
    #niveles necesarios para ingresar a las opciones
    nivel = [["Admin", "No", "Admin"], ["No", "Admin","No"]]
    #tipo de tarea a realizar
    ind=[[1,2,3],[4,5,6]]
    """
    Configurar el comportamiento de un numero de filas
    y columnas para que se expandan con la pantalla
    """
    for i in range(2):
        window.rowconfigure(i, weight=1)
        for j in range(3):  #ciclo para columnas
            window.columnconfigure(j, weight=1)

    """
    ACLARACION:
    NO SE COLOCARON LOS BOTONES MEDIANTE CICLOS DEBIDO A
    QUE PYTHON PIERDE LAS REFERENCIAS DE LAS OPCIONES
    Y FUNCIONES DE CADA BOTON Y ENVIA SIEMPRE LA ULTIMA
    """
    #widget marco vacio
    frame = Frame(window)
    #colocar el widget en 0,0
    frame.grid(row=0, column=0)
    #boton que lleva a la funcion nivel de permiso
    boton1=Button(master=frame, fg="green", text=opc[0][0], width=int(anc / 4), height=int(alt / 2), font=fontStyle,anchor="n", command=lambda: [nivelPermiso(nivel[0][0], userAct, funciones[0][0], window, ind[0][0])])
    #mostrar el boton
    boton1.pack(fill=BOTH, side=TOP, expand=True)

    frame = Frame(window)
    frame.grid(row=0, column=1)
    boton2 = Button(master=frame, fg="green", text=opc[0][1], width=int(anc / 4), height=int(alt / 2), font=fontStyle,anchor="n",command=lambda: [nivelPermiso(nivel[0][1], userAct, funciones[0][1], window, ind[0][1])])
    boton2.pack(fill=BOTH, side=TOP, expand=True)

    frame = Frame(window)
    frame.grid(row=0, column=2)
    boton3 = Button(master=frame, fg="green", text=opc[0][2], width=int(anc / 4), height=int(alt / 2), font=fontStyle,anchor="n",command=lambda: [nivelPermiso(nivel[0][2], userAct, funciones[0][2], window, ind[0][2])])
    boton3.pack(fill=BOTH, side=TOP, expand=True)

    frame = Frame(window)
    frame.grid(row=1, column=0)
    boton5 = Button(master=frame, fg="green", text=opc[1][0], width=int(anc / 4), height=int(alt / 2), font=fontStyle,anchor="n",command=lambda: [nivelPermiso(nivel[1][0], userAct, funciones[1][0], window, ind[1][0])])
    boton5.pack(fill=BOTH, side=TOP, expand=True)

    frame = Frame(window)
    frame.grid(row=1, column=1)
    boton6 = Button(master=frame, fg="green", text=opc[1][1], width=int(anc / 4), height=int(alt / 2), font=fontStyle,anchor="n",command=lambda: [nivelPermiso(nivel[1][1], userAct, funciones[1][1], window, ind[1][1])])
    boton6.pack(fill=BOTH, side=TOP, expand=True)

    frame = Frame(window)
    frame.grid(row=1, column=2)
    boton7 = Button(master=frame, fg="green", text=opc[1][2], width=int(anc / 4), height=int(alt / 2), font=fontStyle,anchor="n",command=lambda: [nivelPermiso(nivel[1][2], userAct, funciones[1][2], window, ind[1][2])])
    boton7.pack(fill=BOTH, side=TOP, expand=True)


    #añadir a imgs la imagen
    imgs.append(PhotoImage(file=nomImgs[0][0]))
    frame = Frame(window)
    frame.grid(row=0, column=0)
    # colocar la img sobre el boton correspondiente
    buti = Button(master=frame, image=imgs[(len(imgs) - 1)], command=lambda: [nivelPermiso(nivel[0][0], userAct, funciones[0][0], window, ind[0][0])])
    buti.pack()

    imgs.append(PhotoImage(file=nomImgs[0][1]))
    frame = Frame(window)
    frame.grid(row=0, column=1)
    buti2 = Button(master=frame, image=imgs[(len(imgs) - 1)], command=lambda: [nivelPermiso(nivel[0][1], userAct, funciones[0][1], window, ind[0][1])])
    buti2.pack()

    imgs.append(PhotoImage(file=nomImgs[0][2]))
    frame = Frame(window)
    frame.grid(row=0, column=2)
    buti = Button(master=frame, image=imgs[(len(imgs) - 1)], command=lambda: [nivelPermiso(nivel[0][2], userAct, funciones[0][2], window, ind[0][2])])
    buti.pack()

    imgs.append(PhotoImage(file=nomImgs[1][0]))
    frame = Frame(window)
    frame.grid(row=1, column=0)
    buti = Button(master=frame, image=imgs[(len(imgs) - 1)], command=lambda: [nivelPermiso(nivel[1][0], userAct, funciones[1][0], window, ind[1][0])])
    buti.pack()

    imgs.append(PhotoImage(file=nomImgs[1][1]))
    frame = Frame(window)
    frame.grid(row=1, column=1)
    buti2 = Button(master=frame, image=imgs[(len(imgs) - 1)], command=lambda: [nivelPermiso(nivel[1][1], userAct, funciones[1][1], window, ind[1][1])])
    buti2.pack()

    imgs.append(PhotoImage(file=nomImgs[1][2]))
    frame = Frame(window)
    frame.grid(row=1, column=2)
    buti = Button(master=frame, image=imgs[(len(imgs) - 1)], command=lambda: [nivelPermiso(nivel[1][2], userAct, funciones[1][2], window, ind[1][2])])
    buti.pack()

    frm = Frame(window)
    frm.grid(row=2, column=0)
    #Boton con la funcion de cambiar el usuario
    botChange = Button(master=frm, bg="blue", fg="white", text="Cambio usuario", width=anc, font=atras,command= lambda: cambio_usuario(window,userAct))  # crear boton
    botChange.pack(fill=BOTH)
    frm = Frame(window)
    frm.grid(row=2, column=1)
    #widget que muestra el usuario activo
    Label(master=frm, fg="black", text="Usuario Activo", width=anc, font=atras).pack(fill=BOTH,side=TOP)
    usuario = Label(master=frm, fg="black",textvariable=userAct, width=anc, font=atras)
    usuario.pack(fill=BOTH,side=BOTTOM)
    frm = Frame(window)  # crear widget vacio
    frm.grid(row=2, column=2)
    fecha = date.today()
    #widget que contiene la fecha actual
    Label(master=frm,text=fecha, fg="black", width=anc, font=atras).pack(fill=BOTH,side=TOP)
    #crea el label que contendra el reloj
    myClock = Label(master=frm, fg="black", width=anc, font=atras)
    #asigna al reloj una hora inicial cualquiera
    myClock['text'] = '21:18:00'
    #muestra el reloj
    myClock.pack(fill=BOTH,side=BOTTOM)
    strftime('%H:%M:%S')
    """
    Funciones que controlan el tiempo de
    actualizacion del reloj
    """
    def tic():
        myClock['text'] = strftime('%H:%M:%S')
    tic()
    def tac():
        tic()
        myClock.after(1000, tac)
    tac()
    window.mainloop()
"""
Funcion intermedia para ingresar productos
ya sea un nuevo producto o mas de uno
ya existente
"""
def subM_ingres(pant):
    anc,alt=tam_pantalla() #Definir tamaño de pantalla
    ventana1 = Toplevel()  #Crear Ventana
    nomImg = ["nuevo.png", "existente.png"]
    ventana1.attributes('-fullscreen',True) #Ampliar ventana a modo pantalla completa
    atras = fuente_salida()
    estilo = tkFont.Font(family="ALGERIAN", size=int(alt/12))  #Cambiar tipo de fuente y tamaño
    opc2=["Nuevo","Existente"]  #lista de opciones
    ventana1.rowconfigure(0, weight=1)
    ventana1.columnconfigure(0, weight=1)  # ajustar tamaño de las columnas
    frm = Frame(ventana1)  # crear widget vacio
    frm.grid(row=0, column=0)  # posicionar frame en i fila y j columna
    bot = Button(master=frm, fg="red", text=opc2[0], width=int(anc / 4), height=int(alt / 2), font=estilo,anchor="n",command=lambda:nuevo(ventana1))  # crear boton
    bot.pack(fill=BOTH, side=TOP, expand=True)  # mostrar boton, que se rellena esta arriba y se expande
    ventana1.columnconfigure(1, weight=1)  # ajustar tamaño de las columnas
    frm = Frame(ventana1)  # crear widget vacio
    frm.grid(row=0, column=1)  # posicionar frame en i fila y j columna
    bot = Button(master=frm, fg="red", text=opc2[1], width=int(anc / 4), height=int(alt / 2), font=estilo,anchor="n",command=lambda:existente(ventana1))  # crear boton
    bot.pack(fill=BOTH, side=TOP, expand=True)  # mostrar boton, que se rellena esta arriba y se expande
    for i in range(2):
        refe2 = PhotoImage(file=nomImg[i])
        img.append(PhotoImage(file=nomImg[i]))
        frm = Frame(ventana1)  # crear widget vacio
        frm.grid(row=0, column=i)  # posicionar frame en i fila y j columna
        imgB = Button(master=frm,image=img[(len(img)-1)])
        imgB.pack()
    frm = Frame(ventana1)  # crear widget vacio
    frm.grid(row=1, column=0, columnspan=2)
    botAtras = Button(master=frm, bg="red", fg="white", text="Atras", width=anc, font=atras,command=ventana1.destroy)  # crear boton
    botAtras.pack(fill=BOTH)
    ventana1.transient(master=pant)
    ventana1.grab_set()
    pant.wait_window(ventana1)
"""
Funcion - Menu para añadir mas cantidad de
producto a alguno que ya este presente en
el inventario
"""
def existente(pant):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    l=0
    if(l==0):
        dic = importar()
        l=1
    ventEx = Toplevel()  # Crear Ventana
    ventEx.attributes('-fullscreen', True)  # Ampliar ventana a modo pantalla completa
    for i in range(6):  # ciclo para filas
        ventEx.rowconfigure(i, weight=1)  # ajustar tamaño de las filas
        for j in range(3):  # ciclo para columnas
            ventEx.columnconfigure(j, weight=1)  # ajustar tamaño de las columnas
    fuente = tkFont.Font(family="ALGERIAN", size=int(alt / 30))  # Cambiar tipo de fuente y tamaño
    fuente2 = tkFont.Font(family="ALGERIAN", size=int(alt / 25))  # Cambiar tipo de fuente y tamaño
    atras = fuente_salida()
    frame = Frame(ventEx)  # crear widget vacio
    frame.grid(row=0, column=0,columnspan=3)
    Label(master=frame, text="Ingresar + Producto", font=fuente2, fg="red",width=anc,anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    x={}
    secs=[]
    for i in dic.keys():
        x[i] = []
        secs.append(i)
        for k in dic[i].keys():
            x[i].append(int(k))
    sec = StringVar()
    ids = StringVar()
    fac=StringVar()
    reset=BooleanVar()
    pTot = IntVar(value=0)
    frame = Frame(ventEx)  # crear widget vacio
    Label(frame,text="SECCION DEL PRODUCTO:",font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    menuSec = OptionMenu(frame, sec, *dic.keys())
    menuSec.configure(font=fuente,width=int(anc/25),anchor="w")
    frame1 = Frame(ventEx)  # crear widget vacio
    Label(frame1, text="ID DEL PRODUCTO:", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    menuId = OptionMenu(frame1, ids, '')
    menuId.configure(font=fuente,width=int(anc/25),anchor="w")
    """
    Funcion que actualiza las opciones de
    el optionlist de id segun lo que se
    halla seleccionado en el optionlist de seccion
    """
    def actualizar(*args):
        dats = x[sec.get()]
        ids.set(dats[0])
        menu = menuId['menu']
        menu.delete(0, 'end')
        for dat in dats:
            menu.add_command(label=dat, command=lambda env=dat: ids.set(env))
    sec.trace('w', actualizar)
    nom = StringVar()
    """
    Funcion que actualiza el nombre del id que este
    seleccionado en el correspondiente
    optionlist
    """
    def actualizar2(*args):
        nom.set(dic[sec.get()][int(ids.get())][0])
    ids.trace('w', actualizar2)
    cant=IntVar(value=1)
    udMed=StringVar()
    frame.grid(row=1, column=0,columnspan=3)
    frame1.grid(row=2, column=0,columnspan=3)
    sec.set(secs[0])
    frame3 = Frame(ventEx)  # crear widget vacio
    frame3.grid(row=3, column=0,columnspan=3)
    Label(master=frame3, text="PRODUCTO A AÑADIR:", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    Label(master=frame3, textvariable=nom, font=fuente,width=int(anc/25),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    frame3 = Frame(ventEx)  # crear widget vacio
    frame3.grid(row=4, column=0,columnspan=3)
    Label(master=frame3, text="CANTIDAD DEL PRODUCTO:", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    cantidad = Spinbox(master=frame3, textvariable=cant, from_=1, to=999, font=fuente,width=int(anc / 25))
    cantidad.pack(side=LEFT, fill=BOTH, expand=True)
    frame3 = Frame(ventEx)  # crear widget vacio
    frame3.grid(row=5, column=0, columnspan=3)
    Button(frame3, text="AÑADIR AL INVENTARIO", font=fuente,bg="lightgreen",activebackground="green", width=int(anc / 20),command=lambda: masPro(ventEx,ids.get(),cantidad.get())).pack(side=LEFT, fill=BOTH, expand=True)
    frame3 = Frame(ventEx)  # crear widget vacio
    frame3.grid(row=6, column=0, columnspan=3)
    Button(frame3, text="Atras", font=fuente, bg="red", fg="white", width=int(anc / 20), command=ventEx.destroy).pack(side=LEFT, fill=BOTH, expand=True)
    menuSec.pack(side=LEFT, fill=BOTH, expand=True)
    menuId.pack(side=LEFT, fill=BOTH, expand=True)
    ventEx.transient(master=pant)
    ventEx.grab_set()
    pant.wait_window(ventEx)
"""
Funcion interna que se encarga de añadir
la cantidad de producto extra al
elemento directamente en el excel
prueba.xlsx
"""
def masPro(vt,id,cant):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    ing = Toplevel()  # Crear Ventana
    ing.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 5)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    ing.resizable(0, 0)
    fecha = date.today()
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    ing['bg'] = 'white'
    Label(ing, text="Producto ingresado", bg='white', font=fonte).pack()
    Label(ing, text="", bg='white').pack()
    wb2 = xlrd.open_workbook('prueba.xlsx')
    sheet = wb2.sheet_by_index(0)
    sheet2 = wb2.sheet_by_index(1)
    try:
        dir_cantidad = len(sheet.row_values(0))
    except:
        dir_cantidad = 0
    try:
        dir_cantidad2 = len(sheet.col_values(0))
    except:
        dir_cantidad2 = 1
    try:
        dir_cantidad_2 = len(sheet2.row_values(0))
    except:
        dir_cantidad_2 = 0
    try:
        dir_cantidad2_2 = len(sheet2.col_values(0))
    except:
        dir_cantidad2_2 = 1
    acct = load_workbook('prueba.xlsx')
    hoja = acct.worksheets[0]
    hoja2 = acct.worksheets[1]
    dats=[]
    for i in range(1,dir_cantidad2+1):
        if(str(hoja.cell(row=i,column=2).value)==str(id)):
            dats=[hoja.cell(row=i, column=1).value,int(id),hoja.cell(row=i, column=3).value,int(cant),hoja.cell(row=i, column=5).value,int(cant)*int(hoja.cell(row=i, column=5).value),hoja.cell(row=i, column=8).value,str(fecha)]
            hoja.cell(row=i, column=4).value=int(hoja.cell(row=i, column=4).value)+int(cant)
            break
    for j in range(len(dats)):
        hoja2.cell(row=dir_cantidad2_2+1,column=j+1).value=dats[j]
    acct.save('prueba.xlsx')
    Button(ing, text="Salir", bg='white', font=fonte, command=ing.destroy).pack()
    ing.transient(master=vt)
    ing.grab_set()
    vt.wait_window(ing)
"""
Funcion - Menu que se encarga de añadir
productos nuevos al inventario
"""
def nuevo(ventana1):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    dic=importar()
    ventNew = Toplevel()  # Crear Ventana
    ventNew.attributes('-fullscreen', True)  # Ampliar ventana a modo pantalla completa
    for i in range(9):  #ciclo para filas
        ventNew.rowconfigure(i, weight=1)  #ajustar tamaño de las filas
        for j in range(4):  #ciclo para columnas
            ventNew.columnconfigure(j, weight=1)   #ajustar tamaño de las columnas
    fuente = tkFont.Font(family="ALGERIAN", size=int(alt / 40))  # Cambiar tipo de fuente y tamaño
    fuente2 = tkFont.Font(family="ALGERIAN", size=int(alt / 32))  # Cambiar tipo de fuente y tamaño
    atras = fuente_salida()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=0, column=1, columnspan=2)
    Label(master=frame, text="Ingreso De Productos Nuevos", font=fuente2, fg="red").pack()
    secciones=[]
    opVar=StringVar("")
    for i in dic.keys():
        secciones.append(i)
    secciones.append("Nueva Seccion")
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=1, column=1,columnspan=2)
    Label(master=frame, text="Seleccione la seccion a la que pertenece", font=fuente).pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=2, column=1)
    opVar.set(secciones[0])
    opM = OptionMenu(frame, opVar, *secciones)
    opM.configure(font=fuente)
    opM.pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=2, column=2)
    Label(master=frame, text="Nueva Seccion", font=fuente).pack()
    newSec=Entry(frame,font=fuente,state=DISABLED)
    newSec.pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=3, column=1,columnspan=2)
    Label(master=frame, text="Nombre DEL PRODUCTO:", font=fuente).pack()
    nombre=Entry(master=frame,font=fuente)
    nombre.pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=4, column=1, columnspan=2)
    Label(master=frame, text="CANTIDAD DEL PRODUCTO:", font=fuente).pack()
    cantidad=Spinbox(master=frame, from_=1, to=999,font=fuente)
    cantidad.pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=5, column=1, columnspan=2)
    Label(master=frame, text="PRECIO DE COMPRA (EN $):", font=fuente).pack()
    precioComp = Entry(master=frame, font=fuente)
    precioComp.pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=6, column=1, columnspan=2)
    Label(master=frame, text="GANACIA EXTRA (EN $):", font=fuente).pack()
    ganancia = Entry(master=frame, font=fuente)
    ganancia.pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=7, column=1, columnspan=2)
    Label(master=frame, text="Unidad de media (EJ:lb , kl ,paquete, caja)", font=fuente).pack()
    udMedida = Entry(master=frame, font=fuente)
    udMedida.pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=8, column=1, columnspan=2)
    """
    funcion que habilita o desabilita
    el widget de nueva seccion dependiendo
    de lo seleccionado en el optionlist seccion
    """
    def cambio(*args):
        if (opVar.get() == "Nueva Seccion"):
            newSec.configure(state=NORMAL)
        else:
            newSec.configure(state=DISABLED)
    opVar.trace("w",cambio)
    Button(master=frame,text="AÑADIR",font=fuente,command=lambda :[ingreso_producto(ventNew,dic,opVar.get(),nombre.get(),cantidad.get(),precioComp.get(),ganancia.get(),udMedida.get(),newSec.get())]).pack()
    frame = Frame(ventNew)  # crear widget vacio
    frame.grid(row=9, column=0, columnspan=4)
    Button(master=frame, bg="red", fg="white", text="Atras", width=anc, font=atras,command=ventNew.destroy).pack()
    ventNew.transient(master=ventana1)
    ventNew.grab_set()
    ventana1.wait_window(ventNew)
"""
funcion interna encargada de añadir el producto
nuevo directamente al excel prueba.xlsx
"""
def ingreso_producto(vt,dic_articulos,sec,nom,cant,prc,gan,udm,ns):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    confirmacion = Toplevel()  # Crear Ventana
    confirmacion.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 5)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    confirmacion.resizable(0, 0)
    fecha = date.today()
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    confirmacion['bg'] = 'white'
    Label(confirmacion, text="Producto ingresado", bg='white', font=fonte).pack()
    Label(confirmacion, text="", bg='white').pack()
    prec_venta = int(prc)+int(gan)
    wb2 = xlrd.open_workbook('prueba.xlsx')
    sheet = wb2.sheet_by_index(0)
    sheet2 = wb2.sheet_by_index(1)
    try:
        dir_cantidad = len(sheet.row_values(0))
    except:
        dir_cantidad = 0
    try:
        dir_cantidad2 = len(sheet.col_values(0))
    except:
        dir_cantidad2 = 1
    try:
        dir_cantidad_2 = len(sheet2.row_values(0))
    except:
        dir_cantidad_2 = 0
    try:
        dir_cantidad2_2 = len(sheet2.col_values(0))
    except:
        dir_cantidad2_2 = 1
    acct= load_workbook('prueba.xlsx')
    hoja=acct.worksheets[0]
    hoja2 = acct.worksheets[1]
    cont=""
    if(ns==""):
        Id = Id_productos(dic_articulos,sec)
        dic_articulos[sec][Id] = [nom, int(cant), int(prc), int(gan), prec_venta, udm]
        hoja.cell(row=dir_cantidad2+1, column=1).value = sec
        hoja2.cell(row=dir_cantidad2_2+1, column=1).value = sec
        hoja.cell(row=dir_cantidad2+1, column=2).value = Id
        hoja2.cell(row=dir_cantidad2_2 + 1, column=2).value = Id
        cont=sec
        for j in range(3, dir_cantidad+1):
            hoja.cell(row=dir_cantidad2+1, column=j).value = dic_articulos[sec][Id][j-3]
    else:
        dic_articulos[ns]={}
        Id = Id_productos(dic_articulos, ns)
        dic_articulos[ns][Id] = [nom, int(cant), int(prc), int(gan), prec_venta, udm]
        hoja.cell(row=dir_cantidad2+1,column=1).value=ns
        hoja2.cell(row=dir_cantidad2_2 + 1, column=1).value = ns
        hoja.cell(row=dir_cantidad2+1, column=2).value = Id
        hoja2.cell(row=dir_cantidad2_2 + 1, column=2).value = Id
        cont = ns
        for i in range(3,dir_cantidad+1):
            hoja.cell(row=dir_cantidad2+1, column=i).value = dic_articulos[ns][Id][i-3]
    hoja2.cell(row=dir_cantidad2_2 + 1, column=3).value = dic_articulos[cont][Id][0]
    hoja2.cell(row=dir_cantidad2_2 + 1, column=4).value = dic_articulos[cont][Id][1]
    hoja2.cell(row=dir_cantidad2_2 + 1, column=5).value = dic_articulos[cont][Id][2]
    hoja2.cell(row=dir_cantidad2_2 + 1, column=6).value = int(hoja2.cell(row=dir_cantidad2_2 + 1, column=4).value) * int(hoja2.cell(row=dir_cantidad2_2 + 1, column=5).value)
    hoja2.cell(row=dir_cantidad2_2 + 1, column=7).value = dic_articulos[cont][Id][5]
    hoja2.cell(row=dir_cantidad2_2 + 1, column=8).value = str(fecha)
    acct.save('prueba.xlsx')
    Button(confirmacion,text="Salir",bg='white',font=fonte,command=confirmacion.destroy).pack()
    confirmacion.transient(master=vt)
    confirmacion.grab_set()
    vt.wait_window(confirmacion)
"""
Funcion encargada de registrar
las ventas que se realizen sobre
el inventario
"""
def salida_producto(pant,user):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    l=0
    if(l==0):
        dic = importar()
        l=1
    ventSal = Toplevel()  # Crear Ventana
    ventSal.attributes('-fullscreen', True)  # Ampliar ventana a modo pantalla completa
    for i in range(11):  # ciclo para filas
        ventSal.rowconfigure(i, weight=1)  # ajustar tamaño de las filas
        for j in range(3):  # ciclo para columnas
            ventSal.columnconfigure(j, weight=1)  # ajustar tamaño de las columnas
    fuente = tkFont.Font(family="ALGERIAN", size=int(alt / 30))  # Cambiar tipo de fuente y tamaño
    fuente2 = tkFont.Font(family="ALGERIAN", size=int(alt / 20))  # Cambiar tipo de fuente y tamaño
    atras = fuente_salida()
    frame = Frame(ventSal)  # crear widget vacio
    frame.grid(row=0, column=0,columnspan=3)
    Label(master=frame, text="Venta", font=fuente2, fg="red",width=anc,anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    x={}
    secs=[]
    for i in dic.keys():
        x[i] = []
        secs.append(i)
        for k in dic[i].keys():
            x[i].append(int(k))
    """
    funcion que actualiza los datos a tratar
    cuando se efectua una compra
    """
    def actualizar3(dic):
        dic = importar()
    sec = StringVar()
    ids = StringVar()
    fac=StringVar()
    reset=BooleanVar()
    pTot = IntVar(value=0)
    """
    funcion que determina el nombre
    de la factura de la compra actual
    y que envia a reiniciar el 'carrito'
    asi como el total
    """
    def newFactura(*args):
        lista_archivos = os.listdir()
        name = "fact_"
        for i in range(0, 999999):
            if name + str(i)+".txt" not in lista_archivos:
                factura = name + str(i) + ".txt"
                break
        fac.set(factura)
        pTot.set(0)
        reset.set(True)
    newFactura()
    elementos=[]
    frame = Frame(ventSal)  # crear widget vacio
    Label(frame,text="SECCION DEL PRODUCTO:",font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    menuSec = OptionMenu(frame, sec, *dic.keys())
    menuSec.configure(font=fuente,width=int(anc/25),anchor="w")
    frame1 = Frame(ventSal)  # crear widget vacio
    Label(frame1, text="ID DEL PRODUCTO:", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    menuId = OptionMenu(frame1, ids, '')
    menuId.configure(font=fuente,width=int(anc/25),anchor="w")
    """
    Funcion que actualiza el optionlist de id
    segun lo seleccionado en el optionlist
    de seccion
    """
    def actualizar(*args):
        dats = x[sec.get()]
        ids.set(dats[0])
        menu = menuId['menu']
        menu.delete(0, 'end')
        for dat in dats:
            menu.add_command(label=dat, command=lambda env=dat: ids.set(env))
    sec.trace('w', actualizar)
    nom = StringVar()
    cant=IntVar(value=1)
    price=IntVar()
    total=IntVar()
    max=IntVar()
    udMed=StringVar()
    """
    Funcion que actualiza los datos del producto
    seleccionado con el id
    """
    def actualizar2(*args):
        dic=importar()
        nom.set(dic[sec.get()][int(ids.get())][0])
        max.set(int(dic[sec.get()][int(ids.get())][1]))
        price.set(int(dic[sec.get()][int(ids.get())][4]))
        total.set(cant.get()*price.get())
        udMed.set(dic[sec.get()][int(ids.get())][5])
    ids.trace('w', actualizar2)
    frame.grid(row=1, column=0,columnspan=3)
    frame1.grid(row=2, column=0,columnspan=3)
    sec.set(secs[0])
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=3, column=0,columnspan=3)
    Label(master=frame3, text="PRODUCTO A COMPRAR:", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    Label(master=frame3, textvariable=nom, font=fuente,width=int(anc/25),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=4, column=0,columnspan=3)
    Label(master=frame3, text="CANTIDAD DEL PRODUCTO:", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    cantidad = Spinbox(master=frame3,textvariable=cant, from_=1, to=999, font=fuente,validatecommand=actualizar2,width=int(anc/32))
    cantidad.pack(side=LEFT,fill=BOTH,expand=True)
    cant.trace("w",actualizar2)
    Label(master=frame3, text="Max", font=fuente,width=int(anc/200)).pack(side=BOTTOM,fill=BOTH,expand=True)
    Label(master=frame3, textvariable=max, font=fuente,width=int(anc/200)).pack(side=LEFT,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=5, column=0,columnspan=3)
    Label(master=frame3, text="Precio Unitario", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    Label(master=frame3, textvariable=price, font=fuente,width=int(anc/25),anchor="w").pack(side=BOTTOM,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=6, column=0, columnspan=3)
    Label(master=frame3, text="Precio Productos", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    Label(master=frame3, textvariable=total, font=fuente,width=int(anc/25),anchor="w").pack(side=BOTTOM,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=7, column=0,columnspan=3)
    Label(master=frame3, text="Total a Pagar", font=fuente,width=int(anc/65),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    Label(master=frame3, textvariable=pTot, font=fuente,width=int(anc/25),anchor="w").pack(side=LEFT,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=8, column=0,columnspan=3)
    Button(frame3,text="Añadir al Carrito",font=fuente,width=int(anc/20),anchor="w",command=lambda:addCar()).pack(side=LEFT,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=9, column=0,columnspan=3)
    Button(frame3, text="Eliminar", font=fuente,width=int(anc/20),anchor="w",command=lambda:eliminar()).pack(side=LEFT,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=10, column=0,columnspan=3)
    Button(frame3, text="Finalizar", font=fuente,width=int(anc/20),anchor="w",command=lambda:[finish(elementos,fac.get(),ventSal,user),newFactura(),actualizar3(dic)]).pack(side=LEFT,fill=BOTH,expand=True)
    frame3 = Frame(ventSal)  # crear widget vacio
    frame3.grid(row=11, column=0, columnspan=3)
    Button(frame3, text="Atras", font=fuente,bg="red",fg="white", width=int(anc / 20),command=ventSal.destroy).pack(side=LEFT, fill=BOTH,expand=True)
    """
    Funcion que añade productos a una lista
    que al final sera enviada para registrar la venta
    """
    def addCar(*args):
        #si se intenta comprar mas producto del que hay disponible
        if (cant.get() > max.get()):
            semi = Toplevel()
            semi.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 4)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
            semi.resizable(0, 0)
            fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
            semi['bg'] = '#929297'
            Label(semi, text="Exedida La Cantidad \nDe Producto En Inventario", bg='#929297', font=fonte).pack()
            Button(semi, text="Volver", font=fonte, command=semi.destroy).pack()
            semi.transient(master=ventSal)
            semi.grab_set()
            ventSal.wait_window(semi)
        else:
            if(reset.get()==True):
                del elementos[:]
                elementos.append([ids.get(), nom.get(), cant.get(), udMed.get(), price.get(), total.get()])
                pTot.set(pTot.get() + total.get())
                reset.set(False)
            else:
                elementos.append([ids.get(), nom.get(), cant.get(), udMed.get(), price.get(), total.get()])
                pTot.set(pTot.get() + total.get())
    """
    Funcion encargada de eliminar productos
    del carrito antes de efectuar la compra
    """
    def eliminar(*args):
        sub0=[]
        for m in range(len(elementos)):
            sub0.append(elementos[m][0])
        if(ids.get() not in sub0):
            semi2 = Toplevel()
            semi2.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 4)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
            semi2.resizable(0, 0)
            fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
            semi2['bg'] = '#929297'
            Label(semi2, text="Este Producto No \nSe Ha Añadido", bg='#929297', font=fonte).pack()
            Button(semi2, text="Volver", font=fonte, command=semi2.destroy).pack()
            semi2.transient(master=ventSal)
            semi2.grab_set()
            ventSal.wait_window(semi2)
        else:
            for v in range(len(elementos)):
                if(ids.get()==elementos[v][0]):
                    pTot.set(pTot.get() - elementos[v][5])
                    del elementos[v]
                    break
    menuSec.pack(side=LEFT,fill=BOTH,expand=True)
    menuId.pack(side=LEFT,fill=BOTH,expand=True)
    ventSal.transient(master=pant)
    ventSal.grab_set()
    pant.wait_window(ventSal)

"""
Funcion encargada de restar lo comprado al inventario
asi como mostrar un preview del recibo que
se puede o no imprimir
"""
def finish(arti,nomFac,pant,user):
    hoy=date.today()
    hora= datetime.now()
    dat=str(hoy)+" - "+str(hora.hour)+":"+str(hora.minute)+":"+str(hora.second)
    total=0
    wb2 = xlrd.open_workbook('prueba.xlsx')
    sheet = wb2.sheet_by_index(0)
    sheet2 = wb2.sheet_by_index(2)
    for i in range(len(arti)):
        total=total+arti[i][5]
    try:
        dir_cantidad = len(sheet.row_values(0))
    except:
        dir_cantidad = 0
    try:
        dir_cantidad2 = len(sheet.col_values(0))
    except:
        dir_cantidad2 = 1
    try:
        dir_cantidad_2 = len(sheet2.row_values(0))
    except:
        dir_cantidad_2 = 0
    try:
        dir_cantidad2_2 = len(sheet2.col_values(0))
    except:
        dir_cantidad2_2 = 1
    acct = load_workbook('prueba.xlsx')
    hoja = acct.worksheets[0]
    hoja2 = acct.worksheets[2]
    dats2=[]
    for r in range(len(arti)):
        for u in range(1, dir_cantidad2+1):
            if (hoja.cell(row=u, column=2).value == int(arti[r][0])):
                hoja.cell(row=u, column=4).value = hoja.cell(row=u, column=4).value - arti[r][2]
                dats2.append([hoja.cell(row=u, column=1).value,int(arti[r][0]),arti[r][1],int(arti[r][2]),int(arti[r][4]),arti[r][3],int(hoja.cell(row=u, column=6).value),(int(hoja.cell(row=u, column=6).value)*int(arti[r][2])),nomFac[:-4],user])
                break
    for g in range(len(dats2)):
        for b in range(1,len(dats2[0])+1):
            hoja2.cell(row=dir_cantidad2_2+1+g, column=b).value = dats2[g][b-1]
    acct.save('prueba.xlsx')
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    venFinal=Toplevel()
    venFinal.geometry(str(int(anc / 2.2)) + 'x' + str(int(alt/1.5)) + '+' + str(int(anc / 2) - int(anc / 8)) + '+' + str(int(alt / 2) - int(alt / 2)))
    venFinal.resizable(0, 0)
    for q in range(8+len(arti)):  # ciclo para filas
        venFinal.rowconfigure(q, weight=1)  # ajustar tamaño de las filas
    for w in range(6):  # ciclo para columnas
        venFinal.columnconfigure(w, weight=1)  # ajustar tamaño de las columnas
    venFinal['bg'] = '#929297'
    canvas = Canvas(venFinal)
    frame = Frame(canvas)

    vertscroll = Scrollbar(canvas, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=vertscroll.set)

    """
    Funcion encargada de leer los movimientos
    de el boton central del mouse e interpretarkos
    como movimiento del scrollbar
    """
    def on_mouse_scroll(event):
        if event.delta:
            canvas.yview_scroll(-1 * int(event.delta / 120), 'units')
        else:
            canvas.yview_scroll(1 if event.num == 5 else -1, 'units')

    venFinal.bind('<Configure>', lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
    venFinal.bind('<MouseWheel>', lambda event: on_mouse_scroll(event))
    venFinal.bind('<Button-4>', lambda event: on_mouse_scroll(event))
    venFinal.bind('<Button-5>', lambda event: on_mouse_scroll(event))

    canvas.pack(side=LEFT, fill=BOTH, expand=1)
    canvas.create_window((0, 0), window=frame, anchor="nw")
    vertscroll.pack(side=RIGHT, fill=Y)

    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño

    Label(frame, text=nomFac, font=fonte,width=int(anc / 60)).grid(row=0,column=0,columnspan=6)
    Label(frame, text=dat, font=fonte, width=int(anc / 60)).grid(row=1,column=0,columnspan=6)
    Label(frame, text="Cajero:"+user, font=fonte, width=int(anc / 60)).grid(row=2, column=0, columnspan=6)
    factu=open(nomFac,"w")
    tx="Cajero:"+user
    factu.write(dat.center(53, " ")+"\n")
    factu.write(tx.center(53, " ") + "\n")
    factu.write("ID".ljust(5," ")+"|NOMBRE".ljust(15," ")+"|CANT.".ljust(7," ")+"|UD.MEDIDA".ljust(10," ")+"|$ X UD".ljust(8," ")+"|$ X ALL".ljust(8," ")+"\n")
    for z in range(len(arti)):
        if(len(arti[z][1])>13):
            factu.write(str(arti[z][0]).ljust(5," ")+"|"+str(arti[z][1][:14]).ljust(14," ")+"|"+str(arti[z][2]).ljust(6," ")+"|"+str(arti[z][3]).ljust(9," ")+"|"+str(arti[z][4]).ljust(7, " ")+"|"+str(arti[z][5]).ljust(7," ")+"\n")
        else:
            factu.write(str(arti[z][0]).ljust(5," ")+"|"+str(arti[z][1]).ljust(14," ")+"|"+str(arti[z][2]).ljust(6," ")+"|"+str(arti[z][3]).ljust(9," ")+"|"+str(arti[z][4]).ljust(7," ")+"|"+str(arti[z][5]).ljust(7," ")+"\n")
    factu.write("Total:$"+str(total).center(53, " ") + "\n")
    factu.close()
    for j in range(len(arti)+1):
        if(j==0):
            Label(frame, text="ID", font=fonte, anchor="w", width=int(anc / 270)).grid(row=j + 3, column=0)
            Label(frame, text="NOMBRE", font=fonte, anchor="w", width=int(anc / 70)).grid(row=j + 3, column=1)
            Label(frame, text="CANT.", font=fonte, anchor="w", width=int(anc / 260)).grid(row=j + 3, column=2)
            Label(frame, text="U.MED", font=fonte, anchor="w", width=int(anc / 200)).grid(row=j + 3, column=3)
            Label(frame, text="$ X UD.", font=fonte, anchor="w", width=int(anc / 200)).grid(row=j + 3, column=4)
            Label(frame, text="$ X ALL", font=fonte, anchor="w", width=int(anc / 200)).grid(row=j + 3, column=5)
        else:
            Label(frame, text=arti[j-1][0], font=fonte, anchor="w", width=int(anc / 270)).grid(row=j + 3, column=0)
            if len(arti[j-1][1]) > 13:
                Label(frame, text=arti[j-1][1][:int(len(arti[j-1][1]) / 2)] + "\n" + arti[j-1][1][int(len(arti[j-1][1]) / 2):],font=fonte, anchor="w", width=int(anc / 70)).grid(row=j + 3, column=1)
            else:
                Label(frame, text=arti[j-1][1], font=fonte, anchor="w", width=int(anc / 70)).grid(row=j + 3, column=1)
            Label(frame, text=arti[j-1][2], font=fonte, anchor="w", width=int(anc / 260)).grid(row=j + 3, column=2)
            Label(frame, text=arti[j-1][3], font=fonte, anchor="w", width=int(anc / 200)).grid(row=j + 3, column=3)
            Label(frame, text=arti[j-1][4], font=fonte, anchor="w", width=int(anc / 200)).grid(row=j + 3, column=4)
            Label(frame, text=arti[j-1][5], font=fonte, anchor="w", width=int(anc / 200)).grid(row=j + 3, column=5)
    Label(frame, text="Total a pagar", font=fonte).grid(row=j + 4, column=0, columnspan=6)
    Label(frame, text=total, font=fonte).grid(row=j + 5, column=0, columnspan=6)
    Button(frame, text="Imprimir Factura", font=fonte, command=lambda :[imprimir(arti,dat,total)]).grid(row=j+6,column=0,columnspan=6)
    Button(frame, text="Salir", font=fonte, command=venFinal.destroy).grid(row=j+7,column=0,columnspan=6)
    venFinal.transient(master=pant)
    venFinal.grab_set()
    pant.wait_window(venFinal)
"""
Funcion encargada de enviar a la impresora
los datos del recibo que se pidio
imprimir

NOTA: DEFECTUOSA PERO FUNCIONAL
"""
def imprimir(arti,dat,tot):
    from win32printing import Printer
    with Printer(linegap=0) as printer:
        printer.text(dat.center(53, " "), align="left",font_config=None)
        printer.text("ID".ljust(6, " ") + "NOMBRE".ljust(15, " ") + "CANT.".ljust(7, " ") + "UD.MEDIDA".ljust(10," ") + "$ X UD".ljust(8, " ") + "$ X ALL".ljust(8, " "),align="left",font_config=None)
        for z in range(len(arti)):
            if (len(arti[z][1]) > 13):
                printer.text(str(arti[z][0]).ljust(5, " ") +  str(arti[z][1][:14]).ljust(20, " ") +  str(arti[z][2]).ljust(10, " ") + str(arti[z][3]).ljust(15, " ") + str(arti[z][4]).ljust(15," ") + str(arti[z][5]).ljust(15, " "),align="left",font_config=None)
            else:
                printer.text(str(arti[z][0]).ljust(5, " ") + str(arti[z][1]).ljust(20, " ") + str(arti[z][2]).ljust(10, " ") + str(arti[z][3]).ljust(15, " ") + str(arti[z][4]).ljust(15, " ") + str(arti[z][5]).ljust(15, " "),font_config=None)
        printer.text(tot.center(53, " "), align="left", font_config=None)
"""
Funcion encargada de mostrar en una tabla
el inventario que tiene la tienda
"""
def inventario(pant):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    dic = importar2()
    inv = Toplevel()  # Crear Ventana
    inv.attributes('-fullscreen', True)  # Ampliar ventana a modo pantalla completa

    canvas = Canvas(inv)
    frame = Frame(canvas)

    vertscroll = Scrollbar(canvas, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=vertscroll.set)

    """
    Funcion que detecta cambios de la rueda del raton
    para mover el scrollbar
    """
    def on_mouse_scroll(event):
        if event.delta:
            canvas.yview_scroll(-1 * int(event.delta / 120), 'units')
        else:
            canvas.yview_scroll(1 if event.num == 5 else -1, 'units')

    inv.bind('<Configure>', lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
    inv.bind('<MouseWheel>', lambda event: on_mouse_scroll(event))
    inv.bind('<Button-4>', lambda event: on_mouse_scroll(event))
    inv.bind('<Button-5>', lambda event: on_mouse_scroll(event))

    canvas.pack(side=LEFT, fill=BOTH, expand=1)
    canvas.create_window((0, 0), window=frame, anchor="nw")
    vertscroll.pack(side=RIGHT, fill=Y)

    fuente = tkFont.Font(family="ALGERIAN", size=int(alt / 25))  # Cambiar tipo de fuente y tamaño
    fuente2 = tkFont.Font(family="ALGERIAN", size=int(alt / 20))  # Cambiar tipo de fuente y tamaño
    atras = fuente_salida()
    wb = xlrd.open_workbook('prueba.xlsx')
    sheet = wb.sheet_by_index(0)
    dir_cantidad = len(sheet.row_values(0))
    dir_cantidad2 = len(sheet.col_values(0))
    Label(frame, text="Inventario", font=fuente2,fg="green",width=int(anc/31), relief="groove", bd=5).grid(row=0, column=0,columnspan=8)
    x=[]
    for j in range(0,5):
        inv.columnconfigure(j, weight=1)
    for i in range(1, dir_cantidad2+1):
        x.append([])
        for y in range(0, dir_cantidad):
            x[i-1].append(sheet.cell(i-1, y).value)
    for t in range(len(x)):
        del x[t][4:6]
    for k in range(len(x)):
        if(k==0):
            Label(frame, width=int(anc / 110), text=x[k][0], font=fuente,relief="groove", bd=5).grid(row=k + 1, column=0)
            Label(frame, width=int(anc / 300), text=x[k][1], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=1)
            Label(frame, width=int(anc / 80), text=x[k][2], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=2)
            Label(frame, width=int(anc / 260), text=x[k][3][:4], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=3)
            Label(frame, width=int(anc / 200), text=x[k][4], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=4)
            Label(frame, width=int(anc / 160), text=x[k][5][:6], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=5)
        else:
            if(len(x[k][0])>10):
                Label(frame, width=int(anc / 110), text=x[k][0][:10], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=0)
            else:
                Label(frame, width=int(anc / 110), text=x[k][0], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=0)
            Label(frame, width=int(anc / 300), text=int(x[k][1]), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=1)
            if (len(x[k][2]) > 20):
                Label(frame, width=int(anc / 80), text=x[k][2][:20], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=2)
            else:
                Label(frame, width=int(anc / 80), text=x[k][2], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=2)
            Label(frame, width=int(anc / 260), text=int(x[k][3]), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=3)
            Label(frame, width=int(anc / 200), text=int(x[k][4]), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=4)
            Label(frame, width=int(anc / 160), text=x[k][5], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=5)
    Button(frame, text="Salir",width=int(anc/31), font=fuente2, fg="Red",command=inv.destroy).grid(row=k+2, column=0, columnspan=8)
    inv.transient(master=pant)
    inv.grab_set()
    pant.wait_window(inv)

"""
Funcion encargada de mostrar en una tabla los
gastos hechos por la tienda al ingresar
productos ya sean nuevos o ya existentes
"""
def gastos(pant):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    dic = importar2()
    gast = Toplevel()  # Crear Ventana
    gast.attributes('-fullscreen', True)  # Ampliar ventana a modo pantalla completa

    canvas = Canvas(gast)
    frame = Frame(canvas)

    vertscroll = Scrollbar(canvas, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=vertscroll.set)

    def on_mouse_scroll(event):
        if event.delta:
            canvas.yview_scroll(-1 * int(event.delta / 120), 'units')
        else:
            canvas.yview_scroll(1 if event.num == 5 else -1, 'units')

    gast.bind('<Configure>', lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
    gast.bind('<MouseWheel>', lambda event: on_mouse_scroll(event))
    gast.bind('<Button-4>', lambda event: on_mouse_scroll(event))
    gast.bind('<Button-5>', lambda event: on_mouse_scroll(event))

    canvas.pack(side=LEFT, fill=BOTH, expand=1)
    canvas.create_window((0, 0), window=frame, anchor="nw")
    vertscroll.pack(side=RIGHT, fill=Y)

    fuente = tkFont.Font(family="ALGERIAN", size=int(alt / 35))  # Cambiar tipo de fuente y tamaño
    fuente2 = tkFont.Font(family="ALGERIAN", size=int(alt / 20))  # Cambiar tipo de fuente y tamaño
    atras = fuente_salida()
    wb = xlrd.open_workbook('prueba.xlsx')
    sheet = wb.sheet_by_index(1)
    dir_cantidad = len(sheet.row_values(0))
    dir_cantidad2 = len(sheet.col_values(0))
    Label(frame, text="Gastos", font=fuente2,fg="green",width=int(anc/31.5), relief="groove", bd=5).grid(row=0, column=0,columnspan=8)
    x=[]
    for j in range(0,8):
        gast.columnconfigure(j, weight=1)
    for i in range(1, dir_cantidad2+1):
        x.append([])
        for y in range(0, dir_cantidad):
            x[i-1].append(sheet.cell(i-1, y).value)
    total=0
    for k in range(len(x)):
        if(k==0):
            Label(frame, width=int(anc / 150), text=x[k][7], font=fuente,relief="groove", bd=5).grid(row=k + 1, column=0)
            Label(frame, width=int(anc / 140), text=x[k][0], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=1)
            Label(frame, width=int(anc / 300), text=x[k][1], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=2)
            Label(frame, width=int(anc / 95), text=x[k][2], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=3)
            Label(frame, width=int(anc / 250), text=x[k][3][:4], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=4)
            Label(frame, width=int(anc / 130), text="P.Compra", font=fuente, relief="groove", bd=5).grid(row=k + 1, column=5)
            Label(frame, width=int(anc / 100), text="Gasto Total", font=fuente, relief="groove", bd=5).grid(row=k + 1,column=6)
            Label(frame, width=int(anc / 160), text=x[k][6], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=7)
        else:
            total=total+int(x[k][5])
            Label(frame, width=int(anc / 150), text=x[k][7], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=0)
            if(len(x[k][0])>10):
                Label(frame, width=int(anc / 140), text=x[k][0][:10], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=1)
            else:
                Label(frame, width=int(anc / 140), text=x[k][0], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=1)
            Label(frame, width=int(anc / 300), text=int(x[k][1]), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=2)
            if (len(x[k][2]) > 15):
                Label(frame, width=int(anc / 95), text=x[k][2][:15], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=3)
            else:
                Label(frame, width=int(anc / 95), text=x[k][2], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=3)
            Label(frame, width=int(anc / 250), text=int(x[k][3]), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=4)
            Label(frame, width=int(anc / 130), text="$"+str(int(x[k][4])), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=5)
            Label(frame, width=int(anc / 100), text="$"+str(int(x[k][5])), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=6)
            Label(frame, width=int(anc / 160), text=x[k][6], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=7)
    Label(frame, width=int(anc/35), text="Gastos Totales", font=fuente, relief="groove", bd=5).grid(row=k + 2, column=0,columnspan=4)
    Label(frame, width=int(anc / 35), text="$"+str(total), font=fuente, relief="groove", bd=5).grid(row=k + 2,column=4,columnspan=4)
    Button(frame, text="Salir",width=int(anc/31.5), font=fuente2, fg="Red",command=gast.destroy).grid(row=k+3, column=0, columnspan=8)
    gast.transient(master=pant)
    gast.grab_set()
    pant.wait_window(gast)
"""
Funcion encargada de mostrar las ganancias
de todos las ventas realizadas hasta la fecha
en una tabla
"""
def ganancias(pant):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    dic = importar3()
    ganc = Toplevel()  # Crear Ventana
    ganc.attributes('-fullscreen', True)  # Ampliar ventana a modo pantalla completa

    canvas = Canvas(ganc)
    frame = Frame(canvas)

    vertscroll = Scrollbar(canvas, orient='vertical', command=canvas.yview)
    canvas.configure(yscrollcommand=vertscroll.set)

    def on_mouse_scroll(event):
        if event.delta:
            canvas.yview_scroll(-1 * int(event.delta / 120), 'units')
        else:
            canvas.yview_scroll(1 if event.num == 5 else -1, 'units')

    ganc.bind('<Configure>', lambda _: canvas.configure(scrollregion=canvas.bbox("all")))
    ganc.bind('<MouseWheel>', lambda event: on_mouse_scroll(event))
    ganc.bind('<Button-4>', lambda event: on_mouse_scroll(event))
    ganc.bind('<Button-5>', lambda event: on_mouse_scroll(event))

    canvas.pack(side=LEFT, fill=BOTH, expand=1)
    canvas.create_window((0, 0), window=frame, anchor="nw")
    vertscroll.pack(side=RIGHT, fill=Y)

    fuente = tkFont.Font(family="ALGERIAN", size=int(alt / 40))  # Cambiar tipo de fuente y tamaño
    fuente2 = tkFont.Font(family="ALGERIAN", size=int(alt / 30))  # Cambiar tipo de fuente y tamaño
    atras = fuente_salida()
    wb = xlrd.open_workbook('prueba.xlsx')
    sheet = wb.sheet_by_index(2)
    dir_cantidad = len(sheet.row_values(0))
    dir_cantidad2 = len(sheet.col_values(0))
    Label(frame, text="Ganancias", font=fuente2,fg="green",width=int(anc/20.2), relief="groove", bd=5).grid(row=0, column=0,columnspan=10)
    x=[]
    for j in range(0,10):
        ganc.columnconfigure(j, weight=1)
    for i in range(1, dir_cantidad2+1):
        x.append([])
        for y in range(0, dir_cantidad):
            x[i-1].append(sheet.cell(i-1, y).value)
    total=0
    for k in range(len(x)):
        if(k==0):
            Label(frame, width=int(anc / 160), text=x[k][9], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=0)
            Label(frame, width=int(anc / 150), text=x[k][8], font=fuente,relief="groove", bd=5).grid(row=k + 1, column=1)
            Label(frame, width=int(anc / 150), text=x[k][0], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=2)
            Label(frame, width=int(anc / 300), text=x[k][1], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=3)
            Label(frame, width=int(anc / 110), text=x[k][2], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=4)
            Label(frame, width=int(anc / 250), text=x[k][3][:4], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=5)
            Label(frame, width=int(anc / 150), text="Prc.Ud", font=fuente, relief="groove", bd=5).grid(row=k + 1, column=6)
            Label(frame, width=int(anc / 150), text=x[k][5][:6], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=7)
            Label(frame, width=int(anc / 180), text="Gan.Ud", font=fuente, relief="groove", bd=5).grid(row=k + 1,column=8)
            Label(frame, width=int(anc / 150), text="Gan.Tot", font=fuente, relief="groove", bd=5).grid(row=k + 1,column=9)
        else:
            total=total+int(x[k][7])
            Label(frame, width=int(anc / 160), text=x[k][9], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=0)
            Label(frame, width=int(anc / 150), text=x[k][8], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=1)
            if(len(x[k][0])>10):
                Label(frame, width=int(anc / 150), text=x[k][0][:10], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=2)
            else:
                Label(frame, width=int(anc / 150), text=x[k][0], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=2)
            Label(frame, width=int(anc / 300), text=int(x[k][1]), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=3)
            if (len(x[k][2]) > 14):
                Label(frame, width=int(anc / 110), text=x[k][2][:14], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=4)
            else:
                Label(frame, width=int(anc / 110), text=x[k][2], font=fuente, relief="groove", bd=5).grid(row=k + 1,column=4)
            Label(frame, width=int(anc / 250), text=int(x[k][3]), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=5)
            Label(frame, width=int(anc / 150), text="$"+str(int(x[k][4])), font=fuente, relief="groove", bd=5).grid(row=k + 1, column=6)
            Label(frame, width=int(anc / 150), text=x[k][5], font=fuente, relief="groove", bd=5).grid(row=k + 1, column=7)
            Label(frame, width=int(anc / 180), text="$"+str(int(x[k][6])), font=fuente, relief="groove", bd=5).grid(row=k + 1,column=8)
            Label(frame, width=int(anc / 150), text="$"+str(int(x[k][7])), font=fuente, relief="groove", bd=5).grid(row=k + 1,column=9)
    Label(frame, width=int(anc/30), text="Ganancias Netas", font=fuente, relief="groove", bd=5).grid(row=k + 2, column=0,columnspan=5)
    Label(frame, width=int(anc / 31), text="$"+str(total), font=fuente, relief="groove", bd=5).grid(row=k + 2,column=5,columnspan=5)
    Button(frame, text="Salir",width=int(anc/20.2), font=fuente2, fg="Red",command=ganc.destroy).grid(row=k+3, column=0, columnspan=10)
    ganc.transient(master=pant)
    ganc.grab_set()
    pant.wait_window(ganc)
"""
Funcion - Menu que permite
seleccionar si acceder con un
usuario existente o registrar uno nuevo

NOTA: PARA REGISTRAR NUEVOS USUARIOS
HAY QUE SER Admin
"""
def cambio_usuario(window,userAct):
    anc,alt=tam_pantalla() #Definir tamaño de pantalla
    cambio = Toplevel()  # Crear Ventana
    cambio.geometry(str(int(anc / 3)) + 'x' + str(int(alt / 2.5)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 4)))
    cambio.resizable(0,0)
    cambio['bg'] = '#929297'
    atras = fuente_salida()
    existente=Button(cambio,bg='#929297',activebackground='#23DF1D',fg='White',text="Existente",anchor="s",font=atras,command= lambda: acceder(cambio,userAct))
    existente.pack(side=TOP,expand=True)
    nuevo = Button(cambio,bg='#929297',activebackground='#23DF1D',fg='White',text="Nuevo usuario",anchor="s",font=atras,command=lambda: nivelPermiso("Admin",userAct,registro,cambio,0,cambio))
    nuevo.pack(side=TOP,expand=True)
    cambio.transient(master=window)
    cambio.grab_set()
    window.wait_window(cambio)
"""
Funcion encargada de permitir 
el ingreso de usuario y contraseña
para iniciar sesion
o ingresar al menu de cambio de
contraseña
"""
def acceder(cambio,userAct):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    acceder = Toplevel()  # Crear Ventana
    acceder.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 2.5)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    acceder.resizable(0, 0)
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    acceder['bg'] = '#929297'
    Label(acceder,text="Ingrese su usuario y contraseña",bg='#929297',font=fonte).pack()
    Label(acceder,text="",bg='#929297').pack()
    Label(acceder,text="USUARIO",bg='#929297',font=fonte).pack()
    usuario = Entry(acceder)
    usuario.pack()
    Label(acceder, text="",bg='#929297').pack()
    Label(acceder, text="CONTRASEÑA",bg='#929297',font=fonte).pack()
    contras = Entry(acceder)
    contras.pack()
    Label(acceder, text="", bg='#929297').pack()
    Button(acceder, text="Acceder", bg='#929297', font=fonte,command=lambda :[verificar(acceder,usuario.get(),contras.get(),userAct,0,0),usuario.delete(0, END),contras.delete(0, END)]).pack()
    Label(acceder, text="", bg='#929297').pack()
    Button(acceder, text="Cambiar contraseña", bg='#929297', font=fonte,command=lambda :[cambio_pass(acceder),usuario.delete(0, END),contras.delete(0, END)]).pack()
    acceder.transient(master=cambio)
    acceder.grab_set()
    cambio.wait_window(acceder)
"""
Funcion interna encargada de verificar
los datos ingresados tanto para
el inicio de sesion como para el
cambio de contraseña
"""
def verificar(pt,user,passwrd,userAct,passwrd2,fnc):
    lista_archivos = os.listdir()
    if(user in lista_archivos):
        arcUs=open(user,"r")
        verPass = arcUs.readlines()
        arcUs.close()
        if (passwrd in verPass[1]):
            if(fnc==0):
                userAct.set(user)
            elif(fnc==1):
                arcUs=open(user,"w")
                arcUs.write(user + "\n")
                arcUs.write(passwrd2)
                arcUs.close()
        else:
            noClave(pt)
    else:
        noUsuario(pt)
"""
Funcion encargada de tomar
el nombre para crear un
usuario nuevo
"""
def registro(cambio):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    regis = Toplevel()  # Crear Ventana
    regis.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 3)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    regis.resizable(0, 0)
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    regis['bg'] = '#929297'
    Label(regis,text="Ingrese su usuario",bg='#929297',font=fonte).pack()
    Label(regis,text="",bg='#929297').pack()
    Label(regis,text="USUARIO",bg='#929297',font=fonte).pack()
    usuario = Entry(regis)
    usuario.pack()
    Label(regis, text="", bg='#929297').pack()
    Button(regis, text="Crear usuario", bg='#929297', font=fonte,command=lambda :[register(usuario.get(),regis),usuario.delete(0, END)]).pack()
    regis.transient(master=cambio)
    regis.grab_set()
    cambio.wait_window(regis)
"""
Funcion encargada de tomar los datos
para el cambio de contraseña
"""
def cambio_pass(pant):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    passCh = Toplevel()  # Crear Ventana
    passCh.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 3)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    passCh.resizable(0, 0)
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    passCh['bg'] = '#929297'
    Label(passCh, text="USUARIO", bg='#929297', font=fonte).pack()
    usuario = Entry(passCh)
    usuario.pack()
    Label(passCh, text="", bg='#929297').pack()
    Label(passCh, text="VIEJA CONTRASEÑA", bg='#929297', font=fonte).pack()
    contras = Entry(passCh)
    contras.pack()
    Label(passCh, text="", bg='#929297').pack()
    Label(passCh, text="NUEVA CONTRASEÑA", bg='#929297', font=fonte).pack()
    contras2 = Entry(passCh)
    contras2.pack()
    Label(passCh, text="", bg='#929297').pack()
    Button(passCh, text="GUARDAR CAMBIOS", bg='#929297', font=fonte,command=lambda: [verificar(passCh,usuario.get(),contras.get(),0,contras2.get(),1), usuario.delete(0, END),contras.delete(0, END),contras2.delete(0, END)]).pack()
    passCh.transient(master=pant)
    passCh.grab_set()
    pant.wait_window(passCh)
"""
Funcion encargada de registrar
un nuevo usuario creando un archivo
con los datos
"""
def register(usuario,regis):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    c, s, l = contra1(usuario)
    p = peso(c, s, l)
    ck = str(checksum(p, s, l))
    newUs = open(usuario, "w")
    newUs.write(usuario + "\n")
    newUs.write(ck)
    newUs.close()
    vent = Toplevel()  # Crear Ventana
    vent.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 4)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    vent.resizable(0, 0)
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    vent['bg'] = '#929297'
    Label(vent, text="LA CONTRASEÑA DEL USUARIO ES:", bg='#929297', font=fonte).pack()
    Label(vent, text="", bg='#929297').pack()
    Label(vent, text=ck, bg='#929297', font=fonte).pack()
    Label(vent, text="", bg='#929297').pack()
    Label(vent, text="Se recomienda cambiar su clave en el menu de Acceso*",bg='#929297').pack()
    Label(vent, text="", bg='#929297').pack()
    Button(vent,text="SALIR",command=vent.destroy).pack()
    vent.transient(master=regis)
    vent.grab_set()
    regis.wait_window(vent)
"""
Funcion para obtener
una semilla para la creacion
de la clave
"""
def contra1(user):
    primos = [101, 113, 197, 229, 271, 337, 419, 499, 541, 557, 659, 691, 727, 733, 761, 839, 881, 919, 983, 1033]
    semilla = random.choice(primos)
    limite = 100007
    cadSal = []
    for i in user:
        if (i != ",") and (i != ".") and (i != ";"):
            cadSal.append(i)
    cad = "".join(cadSal)
    caden = cad.split()
    return caden, semilla, limite

"""
Funcion que calcula el peso de
el nombre de usuario
"""
def peso(lst, semi, limi):
    pesos = []
    for i in range(len(lst)):
        pes = 0
        for j in range(len(lst[i])):
            pes = pes + ((ord(lst[i][j]) * (j + 1)) % semi)
        pesos.append(pes)
    return pesos

"""
Funcion que genera la clave
numerica para el nuevo usuario
"""
def checksum(psos, sem, lim):
    check = 0
    for i in range(len(psos)):
        check = (check + psos[i]) * sem
    if (check > lim):
        check = check % lim
    return check

"""
Funcion encargada de verificar el nivel
necesario para entrar a ciertas opciones
del programa
"""
def nivelPermiso(nivel,actual,func,pant,ind,*args):
    if(actual.get()==nivel):
        argu = []
        for i in args:
            argu.append(i)
        if (ind == 0):
            func(argu[0])
        elif (ind == 1):
            func(pant)
        elif(ind==3):
            func(pant)
        elif(ind==5):
            func(pant)
    elif(ind==6):
        func()
    elif(actual.get()=="__"):
        iniciar(pant)
    elif(nivel == "No"):
        if (ind == 2):
            func(pant)
        elif (ind == 4):
            func(pant, actual.get())
        else:
            func(pant)
    else:
        restringido(pant)
"""
Menu que se genera al digitar
un usuario que no existe
"""
def noUsuario(pant):
    wn=Toplevel()
    Label(wn,text="El usuario no esta registrado").pack()
    ex=Button(wn,text="Salir",command=wn.destroy)
    ex.pack()
    wn.transient(master=pant)
    wn.grab_set()
    pant.wait_window(wn)

"""
Menu que se genera al intentar entrar a una
opcion del programa que esta con un nivel
alto de permisos
"""
def restringido(cambio):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    noAcc = Toplevel()  # Crear Ventana
    noAcc.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 5)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    noAcc.resizable(0, 0)
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    noAcc['bg'] = '#929297'
    Label(noAcc, text="Acceso Restringido", bg='#929297', font=fonte).pack()
    Button(noAcc,text="Salir",font=fonte,command=noAcc.destroy).pack()
    noAcc.transient(master=cambio)
    noAcc.grab_set()
    cambio.wait_window(noAcc)
"""
Menu que se genera al digitar mal la clave
"""
def noClave(pant):
    wn1=Toplevel()
    Label(wn1,text="La contraseña es incorrecta, Vuelva a intentar").pack()
    ex=Button(wn1,text="Salir",command=wn1.destroy)
    ex.pack()
    wn1.transient(master=pant)
    wn1.grab_set()
    pant.wait_window(wn1)
"""
Menu que se genera al intentar entrar a cualquier
opcion del programa sin haber iniciado sesion
con algun usuario
"""
def iniciar(cambio):
    anc, alt = tam_pantalla()  # Definir tamaño de pantalla
    noAcc = Toplevel()  # Crear Ventana
    noAcc.geometry(str(int(anc / 4)) + 'x' + str(int(alt / 5)) + '+' + str(int(anc / 2) - int(anc / 6)) + '+' + str(int(alt / 2) - int(alt / 6)))
    noAcc.resizable(0, 0)
    fonte = tkFont.Font(family="Algerian", size=int(alt / 52))  # Cambiar tipo de fuente y tamaño
    noAcc['bg'] = '#929297'
    Label(noAcc, text="Inicie secion \npara empezar a trabajar", bg='#929297', font=fonte).pack()
    Button(noAcc, text="Salir", font=fonte, command=noAcc.destroy).pack()
    noAcc.transient(master=cambio)
    noAcc.grab_set()
    cambio.wait_window(noAcc)
"""
Funcion que suma 1 al ultimo Id de la sección especifica
retornando el Id que es Id anterior +1
"""
def Id_productos(dic_articulos,seccion):
    producttos=list(dic_articulos.keys())
    idx=0
    for i in range(len(producttos)):
        if producttos[i]==seccion:
            idx=i+1
            break
    Ids=list(dic_articulos[seccion].keys())
    if len(Ids)==0:
        Id=float(idx*100+1)
    else:
        Id_anterior=Ids[len(Ids)-1]
        Id=Id_anterior+1
    return Id

def regla_recursión(puntos,t,iteración):
    color = ['dark orange','yellow','lawn green','spring green','dark turquoise',
                'deep sky blue','blue']
    dibijar_rombo(t,puntos,color[iteración])
    if iteración>0:
        regla_recursión([puntos[0],optener_un_medio(puntos[0],puntos[1]),optener_un_medio(puntos[0],puntos[2]),optener_un_medio(puntos[0],puntos[3])],t,iteración-1)
        regla_recursión([optener_un_medio(puntos[0],puntos[2]),optener_un_medio(puntos[1],puntos[2]),puntos[2],optener_un_medio(puntos[2],puntos[3])],t,iteración-1)
        
def optener_un_medio(punt1,punt2):

    return ((punt1[0]+punt2[0]) /2) , ((punt1[1]+punt2[1]) /2)

def dibijar_rombo(t,puntos,color):
    t.speed(1000)
    t.fillcolor(color)
    t.pencolor(color)
    t.up()
    t.goto(puntos[0][0],puntos[0][1])
    t.down()
    t.begin_fill()
    t.up()
    t.goto(puntos[1][0],puntos[1][1])
    t.down()
    t.goto(puntos[2][0],puntos[2][1])
    t.goto(puntos[3][0],puntos[3][1])
    t.goto(puntos[0][0],puntos[0][1])
    t.end_fill()

def letra_E(x,y,l,t):
    t.pensize(4)
    t.fillcolor('red')
    t.pencolor('black')
    puntosE=[[x,y],[x+(10*l),y],[x+(10*l),y-(3*l)],[x+(8*l),y-(3*l)],[x+(8*l),y-(2*l)],[x+(4*l),y-(2*l)],
    [x+(4*l),y-(5*l)],[x+(5*l),y-(5*l)],[x+(5*l),y-(4*l)],[x+(7*l),y-(4*l)],[x+(7*l),y-(8*l)],[x+(5*l),y-(8*l)],
    [x+(5*l),y-(7*l)],[x+(4*l),y-(7*l)],[x+(4*l),y-(10*l)],[x+(8*l),y-(10*l)],[x+(8*l),y-(9*l)],[x+(10*l),y-(9*l)],
    [x+(10*l),y-(12*l)],[x+(0*l),y-(12*l)],[x+(0*l),y-(10*l)],[x+(1*l),y-(10*l)],[x+(1*l),y-(2*l)],[x+(0*l),y-(2*l)],[x,y]]
    t.up()
    t.goto(x,y)
    t.down()
    t.begin_fill()
    for i in range(len(puntosE)):
        t.goto(puntosE[i][0],puntosE[i][1])
    t.end_fill()    
def letra_I(x,y,l,t):
    t.pensize(4)
    t.fillcolor('white')
    t.pencolor('black')
    x+=100
    y-=50
    puntosI=[[x,y],[x+(6*l),y],[x+(6*l),y-(2*l)],[x+(4.6666*l),y-(2*l)],[x+(4.6666*l),y-(10*l)],[x+(6*l),y-(10*l)],
    [x+(6*l),y-(12*l)],[x+(0*l),y-(12*l)],[x+(0*l),y-(10*l)],[x+(1.33333*l),y-(10*l)],[x+(1.33333*l),y-(2*l)],
    [x+(0*l),y-(2*l)],[x,y]]
    t.up()
    t.goto(x,y)
    t.down()
    t.begin_fill()
    for i in range(len(puntosI)):
         t.goto(puntosI[i][0],puntosI[i][1]) 
    t.end_fill()  

def letra_C(x,y,l,t):
    t.pensize(4)
    t.fillcolor('spring green')
    t.pencolor('black')
    puntosC=[[x,y],[x+(6*l),y],[x+(8*l),y-(2*l)],[x+(8*l),y-(4*l)],[x+(5*l),y-(4*l)],[x+(5*l),y-(4*l)],[x+(5*l),y-(3*l)]
    ,[x+(4.5*l),y-(2*l)],[x+(1.5*l),y-(2*l)],[x+(1*l),y-(3*l)],[x+(1*l),y-(9*l)],[x+(1.5*l),y-(10*l)],[x+(4.5*l),y-(10*l)],[x+(5*l),y-(9*l)],
    [x+(5*l),y-(8*l)],[x+(8*l),y-(8*l)],[x+(8*l),y-(10*l)],[x+(6*l),y-(12*l)],[x+(0*l),y-(12*l)],[x+(-2*l),y-(10*l)],
    [x+(-2*l),y-(2*l)],[x,y]]
    t.up()
    t.goto(x,y)
    t.down()
    t.begin_fill()
    for i in range(len(puntosC)):
         t.goto(puntosC[i][0],puntosC[i][1]) 
    t.end_fill()    

def main_logo():
    #Variable ventana
    wn= turtle.Screen()
    wn.bgcolor("black")
    wn.screensize(1500,1000)
    wn.setup(1500, 1000, 0, 0)
    puntos=[[-450,0],[0,350],[450,0],[0,-350]]
    #Config Tortugas
    t=turtle.Turtle()
    regla_recursión(puntos,t,6)
    letra_E(-90,250,10,t)
    letra_I(-90,250,10,t)
    letra_C(-25,-150,10,t)
    time.sleep(5)
    wn.clear()
    wn.bgcolor("black")
    binevenido()
    wn.exitonclick()
def binevenido():
    b=turtle.Turtle()
    b.hideturtle()
    b.speed(10)
    b.color("yellow")
    b.up()
    b.goto(-700, 150)
    b.down()
    b.right(90)
    b.forward(260)
    b.left(90)
    b.forward(50)
    b.circle(75, 180)
    b.left(180)
    b.circle(55, 180)
    b.forward(50)
    b.up()
    b.goto(-680, 130)
    b.down()
    b.left(90)
    b.forward(70)
    b.left(90)
    b.forward(25)
    b.circle(35, 180)
    b.forward(25)
    b.up()
    b.goto(-680, 20)
    b.down()
    b.left(90)
    b.forward(110)
    b.left(90)
    b.forward(30)
    b.circle(55, 180)
    b.forward(30)

    i=turtle.Turtle()
    i.color("yellow")
    i.speed(10)
    i.hideturtle()
    i.up()
    i.goto(-565, 150)
    i.down()
    i.right(90)
    i.forward(260)
    i.left(90)
    i.forward(30)
    i.left(90)
    i.forward(260)
    i.left(90)
    i.forward(30)

    e=turtle.Turtle()
    e.speed(10)
    e.color("yellow")
    e.hideturtle()
    e.up()
    e.goto(-520, 150)
    e.down()
    e.right(90)
    e.forward(260)
    e.left(90)
    e.forward(100)
    e.left(90)
    e.forward(30)
    e.left(90)
    e.forward(70)
    e.right(90)
    e.forward(85)
    e.right(90)
    e.forward(70)
    e.left(90)
    e.forward(30)
    e.left(90)
    e.forward(70)
    e.right(90)
    e.forward(85)
    e.right(90)
    e.forward(70)
    e.left(90)
    e.forward(30)
    e.left(90)
    e.forward(100)

    n=turtle.Turtle()
    n.color("yellow")
    n.hideturtle()
    n.speed(10)
    n.up()
    n.goto(-405, 150)
    n.down()
    n.right(90)
    n.forward(260)
    n.left(90)
    n.forward(30)
    n.left(90)
    n.forward(180)
    n.right(150)
    n.forward(205)
    n.right(300)
    n.forward(30)
    n.left(90)
    n.forward(260)
    n.left(90)
    n.forward(30)
    n.left(90)
    n.forward(180)
    n.right(150)
    n.forward(205)
    n.right(300)
    n.forward(30)

    v=turtle.Turtle()
    v.color("blue")
    v.hideturtle()
    v.speed(10)
    v.up()
    v.goto(-225, 150)
    v.down()
    v.right(80)
    v.forward(261)
    v.left(80)
    v.forward(50)
    v.left(80)
    v.forward(261)
    v.right(260)
    v.forward(30)
    v.left(80)
    v.forward(196)
    v.right(80)
    v.forward(15)
    v.right(80)
    v.forward(196)
    v.left(80)
    v.forward(30)

    e=turtle.Turtle()
    e.speed(10)
    e.color("blue")
    e.hideturtle()
    e.up()
    e.goto(-75, 150)
    e.down()
    e.right(90)
    e.forward(260)
    e.left(90)
    e.forward(100)
    e.left(90)
    e.forward(30)
    e.left(90)
    e.forward(70)
    e.right(90)
    e.forward(85)
    e.right(90)
    e.forward(70)
    e.left(90)
    e.forward(30)
    e.left(90)
    e.forward(70)
    e.right(90)
    e.forward(85)
    e.right(90)
    e.forward(70)
    e.left(90)
    e.forward(30)
    e.left(90)
    e.forward(100)


    n=turtle.Turtle()
    n.color("blue")
    n.hideturtle()
    n.speed(10)
    n.up()
    n.goto(40, 150)
    n.down()
    n.right(90)
    n.forward(260)
    n.left(90)
    n.forward(30)
    n.left(90)
    n.forward(180)
    n.right(150)
    n.forward(205)
    n.right(300)
    n.forward(30)
    n.left(90)
    n.forward(260)
    n.left(90)
    n.forward(30)
    n.left(90)
    n.forward(180)
    n.right(150)
    n.forward(205)
    n.right(300)
    n.forward(30)


    i=turtle.Turtle()
    i.color("red")
    i.speed(10)
    i.hideturtle()
    i.up()
    i.goto(220, 150)
    i.down()
    i.right(90)
    i.forward(260)
    i.left(90)
    i.forward(30)
    i.left(90)
    i.forward(260)
    i.left(90)
    i.forward(30)

    d=turtle.Turtle()
    d.color("red")
    d.speed(10)
    d.hideturtle()
    d.up()
    d.goto(265, 150)
    d.down()
    d.right(90)
    d.forward(260)
    d.left(90)
    d.forward(30)
    d.circle(130, 180)
    d.forward(30)
    d.up()
    d.goto(295, 120)
    d.down()
    d.left(90)
    d.forward(200)
    d.left(90)
    d.circle(100, 180)

    o=turtle.Turtle()
    o.color("red")
    o.speed(10)
    o.hideturtle()
    o.up()
    o.goto(565 ,-110)
    o.down()
    o.circle(130)
    o.up()
    o.goto(565, -80)
    o.down()
    o.circle(100)
def main():
    main_logo()
    menu()       
main()


